#!/usr/bin/env python3
"""
Holiday Records comparison tool.

Reads an Excel workbook where each TAB (sheet) represents a system
(e.g. WS, D3, Barracuda). Every tab lists currencies (CAD / USD / HKD ...)
and, for each currency, one or more holiday dates in MM/DD/YYYY format.

You nominate one tab as the SOURCE system. The script then builds the union
of every (currency, date) pair seen across all systems and, for each pair,
reports Y/N for whether that pair exists in each system. This surfaces both:

  * dates the source has that another system is missing (source=Y, other=N)
  * dates another system has that the source does not (source=N, other=Y)

Output is a CSV report.

Supported tab layouts (auto-detected per sheet):

  1. LONG  - a "Currency" column and a "Date"/"Holiday" column, many rows:
                Currency | Holiday Date
                CAD      | 01/01/2020
                CAD      | 07/01/2020
                USD      | 01/01/2020

  2. WIDE  - currency codes as column headers, dates listed down each column:
                CAD        | USD        | HKD
                01/01/2020 | 01/01/2020 | 01/01/2020
                07/01/2020 |            | 10/01/2020

Usage:
    python compare_holidays.py --source WS
    python compare_holidays.py --file input/holidays.xlsx --source WS --output output/report.csv
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import glob
import os
import re
import sys
from collections import defaultdict

from openpyxl import load_workbook

# A cell counts as a "currency code" if it is 3 alphabetic characters.
CURRENCY_RE = re.compile(r"^[A-Za-z]{3}$")

# Header keywords used to recognise the LONG layout. The "System" column is
# intentionally NOT listed here — the system name always comes from the tab name.
CURRENCY_HEADERS = {"currency", "ccy", "curr", "currency code", "holiday ccy", "holiday currency"}
DATE_HEADERS = {"date", "dates", "holiday", "holiday date", "holidays", "holiday dates", "value"}

# Fallback textual date formats always tried after the user-selected one.
FALLBACK_DATE_FORMATS = ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%y", "%m-%d-%Y")

# Friendly date-format tokens -> strptime/strftime patterns (longest token first).
_FMT_TOKENS = (("YYYY", "%Y"), ("YY", "%y"), ("MM", "%m"), ("DD", "%d"))


def friendly_to_strftime(token: str) -> str:
    """Convert a friendly format like 'MM/DD/YYYY' to '%m/%d/%Y'."""
    result = token.strip()
    for src, dst in _FMT_TOKENS:
        result = result.replace(src, dst)
    return result


def norm_currency(value) -> str | None:
    """Return an upper-cased 3-letter currency code, or None."""
    if value is None:
        return None
    text = str(value).strip()
    if CURRENCY_RE.match(text):
        return text.upper()
    return None


def parse_date(value, formats):
    """Return a datetime.date from a cell value, or None if it isn't a date.

    Real Excel date cells arrive as datetime and need no format. Text dates are
    parsed with `formats` (the user-chosen format first, then fallbacks).
    """
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in formats:
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def looks_like_currency(text) -> bool:
    return norm_currency(text) is not None


def _find_header_columns(rows):
    """Locate a header row with currency + date columns (LONG layout).

    Returns (header_row_index, currency_col, date_col) or None.
    """
    for r_idx, row in enumerate(rows[:5]):  # header is near the top
        currency_col = date_col = None
        for c_idx, cell in enumerate(row):
            if cell is None:
                continue
            label = str(cell).strip().lower()
            if label in CURRENCY_HEADERS and currency_col is None:
                currency_col = c_idx
            elif label in DATE_HEADERS and date_col is None:
                date_col = c_idx
        if currency_col is not None and date_col is not None:
            return r_idx, currency_col, date_col
    return None


def parse_sheet(rows, formats):
    """Parse one sheet's rows into a set of (currency, date) pairs.

    Handles LONG and WIDE layouts. `rows` is a list of tuples of cell values.
    The system name is NOT taken from the sheet contents (a "System" column is
    ignored) — callers key results by the tab/sheet name instead.
    """
    pairs: set[tuple[str, dt.date]] = set()
    if not rows:
        return pairs

    header = _find_header_columns(rows)
    if header is not None:
        # --- LONG layout ---
        h_idx, cur_col, date_col = header
        for row in rows[h_idx + 1:]:
            if cur_col >= len(row) or date_col >= len(row):
                continue
            currency = norm_currency(row[cur_col])
            date = parse_date(row[date_col], formats)
            if currency and date:
                pairs.add((currency, date))
        return pairs

    # --- WIDE layout: currency codes as column headers ---
    header_row_idx = None
    for r_idx, row in enumerate(rows[:5]):
        currency_cells = [c for c in row if looks_like_currency(c)]
        if len(currency_cells) >= 1 and len(currency_cells) >= sum(
            1 for c in row if c is not None and str(c).strip()
        ) / 2:
            header_row_idx = r_idx
            break

    if header_row_idx is not None:
        header = rows[header_row_idx]
        col_currency = {
            c_idx: norm_currency(cell)
            for c_idx, cell in enumerate(header)
            if norm_currency(cell)
        }
        for row in rows[header_row_idx + 1:]:
            for c_idx, currency in col_currency.items():
                if c_idx < len(row):
                    date = parse_date(row[c_idx], formats)
                    if date:
                        pairs.add((currency, date))
        return pairs

    # --- Fallback: detect a currency column and any date columns by content ---
    n_cols = max(len(row) for row in rows)
    currency_score = [0] * n_cols
    date_score = [0] * n_cols
    for row in rows:
        for c_idx in range(n_cols):
            if c_idx >= len(row):
                continue
            cell = row[c_idx]
            if looks_like_currency(cell):
                currency_score[c_idx] += 1
            elif parse_date(cell, formats) is not None:
                date_score[c_idx] += 1

    if max(currency_score, default=0) == 0:
        return pairs
    cur_col = currency_score.index(max(currency_score))
    date_cols = [i for i, s in enumerate(date_score) if s > 0 and i != cur_col]
    for row in rows:
        if cur_col >= len(row):
            continue
        currency = norm_currency(row[cur_col])
        if not currency:
            continue
        for dc in date_cols:
            if dc < len(row):
                date = parse_date(row[dc], formats)
                if date:
                    pairs.add((currency, date))
    return pairs


def read_workbook(path, formats):
    """Return {system_name: set((currency, date))} for every sheet.

    System name = sheet/tab name (the in-sheet "System" column is ignored).
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    systems = {}
    order = []
    for sheet in wb.worksheets:
        rows = [tuple(r) for r in sheet.iter_rows(values_only=True)]
        systems[sheet.title] = parse_sheet(rows, formats)
        order.append(sheet.title)
    wb.close()
    return systems, order


def resolve_source(systems, requested):
    """Case-insensitive match of the requested source against sheet names."""
    if requested in systems:
        return requested
    lowered = {name.lower(): name for name in systems}
    if requested.lower() in lowered:
        return lowered[requested.lower()]
    return None


def build_report(systems, order, source, out_format):
    """Return (header_row, data_rows) for the CSV report."""
    other_systems = [s for s in order if s != source]
    system_cols = [source] + other_systems

    # Union of every (currency, date) across all systems.
    all_pairs = set()
    for pairs in systems.values():
        all_pairs |= pairs

    header = ["Currency", "Date"] + system_cols + ["Status", "Missing_In"]
    data_rows = []
    for currency, date in sorted(all_pairs, key=lambda p: (p[0], p[1])):
        flags = {}
        for sysname in system_cols:
            flags[sysname] = "Y" if (currency, date) in systems[sysname] else "N"
        missing_in = [s for s in system_cols if flags[s] == "N"]
        status = "MATCH" if not missing_in else "MISMATCH"
        row = [currency, date.strftime(out_format)]
        row += [flags[s] for s in system_cols]
        row += [status, "; ".join(missing_in)]
        data_rows.append(row)
    return header, data_rows, system_cols


def print_summary(systems, order, source, data_rows, system_cols):
    print(f"\nSource system : {source}")
    print(f"Systems ({len(order)}) : {', '.join(order)}")
    print("\nPairs parsed per system:")
    for name in [source] + [s for s in order if s != source]:
        print(f"  {name:<14} {len(systems[name])}")
    total = len(data_rows)
    matches = sum(1 for r in data_rows if r[-2] == "MATCH")
    mismatches = total - matches
    print(f"\nUnique (currency, date) pairs : {total}")
    print(f"  MATCH (in every system)     : {matches}")
    print(f"  MISMATCH                    : {mismatches}")


def main(argv=None):
    parser = argparse.ArgumentParser(description="Compare holiday dates across systems in an Excel workbook.")
    parser.add_argument("--file", help="Path to the .xlsx file (defaults to the newest file in input/).")
    parser.add_argument("--source", required=True, help="Name of the source system (tab name), e.g. WS.")
    parser.add_argument("--output", default="output/holiday_comparison.csv", help="Path for the CSV report.")
    parser.add_argument(
        "--date-format",
        default="MM/DD/YYYY",
        help="Date format for parsing text dates and for the report, e.g. MM/DD/YYYY, DD/MM/YYYY, YYYY-MM-DD.",
    )
    args = parser.parse_args(argv)

    out_format = friendly_to_strftime(args.date_format)
    # Try the user's format first, then common fallbacks, when parsing text dates.
    parse_formats = [out_format] + [f for f in FALLBACK_DATE_FORMATS if f != out_format]

    path = args.file
    if not path:
        candidates = sorted(
            glob.glob("input/*.xlsx") + glob.glob("input/*.xls"),
            key=os.path.getmtime,
            reverse=True,
        )
        if not candidates:
            parser.error("No --file given and no .xlsx found in input/.")
        path = candidates[0]
        print(f"Using input file: {path}")

    if not os.path.exists(path):
        # Allow a bare filename: look for it inside input/.
        alt = os.path.join("input", path)
        if os.path.exists(alt):
            path = alt
            print(f"Using input file: {path}")
        else:
            parser.error(f"File not found: {path} (also looked in input/)")

    systems, order = read_workbook(path, parse_formats)
    if not systems:
        parser.error("Workbook has no sheets.")

    source = resolve_source(systems, args.source)
    if source is None:
        parser.error(
            f"Source system '{args.source}' not found. Available tabs: {', '.join(order)}"
        )

    header, data_rows, system_cols = build_report(systems, order, source, out_format)

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
    with open(args.output, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(data_rows)

    print_summary(systems, order, source, data_rows, system_cols)
    print(f"\nReport written to: {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
